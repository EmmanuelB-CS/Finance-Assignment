import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.drawing.image import Image as openpyxl_image


# Charger le fichier Excel contenant les données
file_path = r"C:\Users\User\Desktop\Finance assignment\Python_version\EURO STOXX 50 Price + Index Data.xlsx"
df = pd.read_excel(file_path, sheet_name='Price Data', index_col=0)

# Sélectionner les 12 actions avec des données continues
selected_stocks = ["ADIDAS (XET)", "ENEL", "KONINKLIJKE AHOLD DELHAIZE", "BBV.ARGENTARIA", "L AIR LQE.SC.ANYME. POUR L ETUDE ET L EPXTN.",
                       "AIRBUS", "ALLIANZ (XET)", "ANHEUSER-BUSCH INBEV", "ASML HOLDING", "AXA",
                       "BASF (XET)", "BAYER (XET)"]

selected_df = df[selected_stocks]

# Calculer les rendements mensuels
monthly_returns = selected_df.pct_change().dropna()

# Calculer les rendements moyens et écart-types mensuels
average_monthly_returns = (monthly_returns.mean()).round(4)
std_dev_monthly_returns = (monthly_returns.std()).round(4)

# Annualiser les rendements et écart-types
annual_average_returns = average_monthly_returns * 12
annual_std_dev_returns = std_dev_monthly_returns * np.sqrt(12)

# Créer un portefeuille équilibré
equally_weighted_portfolio = monthly_returns.mean(axis=1)

# Calculer les rendements moyens et écart-types du portefeuille
average_portfolio_return = (equally_weighted_portfolio.mean()).round(4)
std_dev_portfolio_return = (equally_weighted_portfolio.std()).round(4)

# Annualiser les rendements et écart-types du portefeuille
annual_average_portfolio_return = average_portfolio_return * 12
annual_std_dev_portfolio_return = std_dev_portfolio_return * np.sqrt(12)

# Créer un scatter plot
plt.figure(figsize=(8, 6))
plt.scatter(annual_std_dev_returns, annual_average_returns, c='blue', label='Actions sélectionnées')
plt.scatter(annual_std_dev_portfolio_return, annual_average_portfolio_return, c='red', label='Portefeuille équilibré')
plt.xlabel('Écart-type annuel')
plt.ylabel('Rendement moyen annuel')
plt.title('Performance des actions et du portefeuille équilibré')
plt.legend()

# Enregistrer le scatter plot dans un tampon de mémoire
scatter_plot_buffer = BytesIO()
plt.savefig(scatter_plot_buffer, format='png')
scatter_plot_buffer.seek(0)

# Présenter les résultats sous forme de DataFrames que l'on convertir en excel
results = {
    "Stocks": selected_stocks,
    "Average Monthly Returns": average_monthly_returns,
    "Standard Deviation (Monthly)": std_dev_monthly_returns,
    "Annual Average Returns": annual_average_returns,
    "Annual Standard Deviation": annual_std_dev_returns
}

portfolio_results = {
    "Portfolio": ["Equally Weighted"],
    "Average Monthly Returns": [average_portfolio_return],
    "Standard Deviation (Monthly)": [std_dev_portfolio_return],
    "Annual Average Returns": [annual_average_portfolio_return],
    "Annual Standard Deviation": [annual_std_dev_portfolio_return]
}

stocks_df = pd.DataFrame(results)
portfolio_df = pd.DataFrame(portfolio_results)

print("Rendements des actions sélectionnées :")
print(stocks_df)
print("\nRendements du portefeuille équilibré :")
print(portfolio_df)


output_path = r"C:\Users\User\Desktop\Finance assignment\csv_from_python_code\dataframes.xlsx"

# Charger le fichier Excel s'il existe, sinon créer un nouveau
try:
    book = load_workbook(output_path)
except FileNotFoundError:
    book = None

with pd.ExcelWriter(output_path) as writer:  
    stocks_df.to_excel(writer, sheet_name='Stocks', index=False)
    portfolio_df.to_excel(writer, sheet_name='Portfolio', index=False)


# # This portion of the code is supposed to add the scatter plot to the excel, but for an unknown reason
#   it does not work, I'll try to fix it 
#     book = load_workbook(output_path)
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#     # Insérer le scatter plot dans la feuille Excel
#     image = openpyxl_image(scatter_plot_buffer)
#     image.anchor = writer.sheets['Stocks']['A1']
#     writer.sheets['Stocks'].add_image(image)

# print("Results were stored with the scatter plot in the Excel file")





